import re
import tkinter as tk
from tkinter import messagebox, filedialog, ttk
import xlrd
import openpyxl
from openpyxl.styles import PatternFill, colors

DONOR_INDEX = 1
NAME_INDEX = 2
STREET_INDEX = 3
CITY_INDEX = 4
STATE_INDEX = 5
ZIP_INDEX = 6
PHONE_INDEX = 7
EMAIL_INDEX = 8
COMPANY_INDEX = 9


def format_phone_number(num):
    # Number is not 10-digit US number, don't format
    if re.match('^[0-9]{10}$', num) is None:
        return num

    # Break the number into pieces and assemble properly
    a = num[:3]
    b = num[3:6]
    c = num[6:]

    return '(' + a + ') ' + b + '-' + c


def unformat_phone_number(num):
    return ''.join(re.split('[^0-9]', num))


class SowersFund(tk.Frame):
    def __init__(self, parent):
        super(SowersFund, self).__init__(parent)

        self.db_path = None

        # Loads the database during read operations only
        self.database = None

        # Which row, if any, is currently loaded
        self.loaded_row = None

        self.winfo_toplevel().title('Sowers Fund')

        self.init_ui()

    def init_ui(self):

        # Donor
        tk.Label(text='Donor').grid(column=0, row=0)
        self.donor_field = ttk.Entry(width=36)
        self.donor_field.config(state=tk.DISABLED)
        self.donor_field.grid(column=1, row=0)

        # Name
        tk.Label(text='Full Name').grid(column=0, row=1)
        self.name_field = ttk.Combobox(width=36)
        self.name_field.bind('<KeyRelease>', lambda ev: self.auto_suggest(self.name_field, NAME_INDEX))
        self.name_field.config(state=tk.DISABLED)
        self.name_field.grid(column=1, row=1)

        self.name_button = tk.Button(text='Look Up Name', command=lambda: self.lookup(self.name_field.get(), NAME_INDEX))
        self.name_button.config(state=tk.DISABLED)
        self.name_button.grid(column=2, row=1)

        # Shipping Street
        tk.Label(text='Shipping Street').grid(column=0, row=2)
        self.street_field = tk.Entry(width=36)
        self.street_field.config(state=tk.DISABLED)
        self.street_field.grid(column=1, row=2)

        # Shipping City
        tk.Label(text='Shipping City').grid(column=0, row=3)
        self.city_field = tk.Entry(width=36)
        self.city_field.config(state=tk.DISABLED)
        self.city_field.grid(column=1, row=3)

        # Shipping State
        tk.Label(text='Shipping State').grid(column=0, row=4)
        self.state_field = tk.Entry(width=36)
        self.state_field.config(state=tk.DISABLED)
        self.state_field.grid(column=1, row=4)

        # Zip
        tk.Label(text='Shipping Zip').grid(column=0, row=5)
        self.zip_field = tk.Entry(width=36)
        self.zip_field.config(state=tk.DISABLED)
        self.zip_field.grid(column=1, row=5)

        # Phone
        tk.Label(text='Phone').grid(column=0, row=6)
        self.phone_field = ttk.Combobox(width=36)
        self.phone_field.bind('<KeyRelease>', lambda ev: self.auto_suggest(self.phone_field, PHONE_INDEX))
        self.phone_field.config(state=tk.DISABLED)
        self.phone_field.grid(column=1, row=6)

        self.phone_button = tk.Button(text='Look Up Phone', command=lambda: self.lookup_phone())
        self.phone_button.config(state=tk.DISABLED)
        self.phone_button.grid(column=2, row=6)

        # Email
        tk.Label(text='Email').grid(column=0, row=7)
        self.email_field = tk.Entry(width=36)
        self.email_field.config(state=tk.DISABLED)
        self.email_field.grid(column=1, row=7)

        self.email_button = tk.Button(text='Look Up Email', command=lambda: self.lookup(self.email_field.get(), EMAIL_INDEX))
        self.email_button.config(state=tk.DISABLED)
        self.email_button.grid(column=2, row=7)

        # Company Name
        tk.Label(text='Company Name').grid(column=0, row=8)
        self.company_field = tk.Entry(width=36)
        self.company_field.config(state=tk.DISABLED)
        self.company_field.grid(column=1, row=8)

        tk.Button(text='Clear Form', command=self.confirm_clear).grid(column=0, row=9)

        self.submit_button = tk.Button(text='Submit', command=self.submit)
        self.submit_button.config(state=tk.DISABLED)
        self.submit_button.grid(column=1, row=9)

        tk.Button(text='Open File', command=self.open_db_file).grid(column=2, row=9)

    def open_db_file(self):
        dialog = filedialog.Open(self, filetypes=[('Excel Files', '*.xlsx')])
        fpath = dialog.show()

        if fpath == '':
            return

        self.db_path = fpath

        # Enable ext fields
        self.donor_field.config(state=tk.NORMAL)
        self.name_field.config(state=tk.NORMAL)
        self.street_field.config(state=tk.NORMAL)
        self.city_field.config(state=tk.NORMAL)
        self.state_field.config(state=tk.NORMAL)
        self.zip_field.config(state=tk.NORMAL)
        self.phone_field.config(state=tk.NORMAL)
        self.email_field.config(state=tk.NORMAL)
        self.company_field.config(state=tk.NORMAL)

        # Enable buttons
        self.name_button.config(state=tk.NORMAL)
        self.phone_button.config(state=tk.NORMAL)
        self.email_button.config(state=tk.NORMAL)
        self.submit_button.config(state=tk.NORMAL)

        messagebox.showinfo('Success', 'Successfully loaded file')

    def auto_suggest(self, field, row_index):
        if len(field.get().strip()) < 3 or type(field) is not ttk.Combobox:
            return

        # Convert value to lowercase and strip whitespace for case-insensitive lookups
        value = field.get().lower().strip()

        suggestions = []

        self.database = xlrd.open_workbook(self.db_path).sheet_by_index(0)

        for row in range(self.database.nrows):
            cell_value = self.database.row(row)[row_index].value

            # Convert floats to ints to remove trailing decimal places
            if type(cell_value) is float:
                cell_value = int(cell_value)

            # Store the original value for display purposes
            cell_value = str(cell_value)

            # Convert to lowercase string and strip whitespace for the item to search
            search_value = cell_value.lower().strip()

            # Read phone numbers back without formatting for lenient comparison
            if row_index == PHONE_INDEX:
                search_value = unformat_phone_number(search_value)

            # Entry value is a str, cell value could be anything, so we need to convert
            if value in search_value:
                suggestions.append(cell_value)

        field['values'] = suggestions

    # Phone needs a separate method due to formatting of 10-digit US numbers
    def lookup_phone(self):
        value = unformat_phone_number(self.phone_field.get())
        print(value)
        self.lookup(value, PHONE_INDEX)

    def lookup(self, value, row_index):
        # Convert value to lowercase and strip whitespace for case-insensitive lookups
        value = value.lower().strip()

        self.database = xlrd.open_workbook(self.db_path).sheet_by_index(0)

        for row in range(self.database.nrows):
            cell_value = self.database.row(row)[row_index].value

            # Convert floats to ints to remove trailing decimal places
            if type(cell_value) is float:
                cell_value = int(cell_value)

            # Convert to lowercase string and strip whitespace
            cell_value = str(cell_value).lower().strip()

            # Read phone numbers back without formatting for lenient comparison
            if row_index == PHONE_INDEX:
                cell_value = unformat_phone_number(cell_value)

            # Entry value is a str, cell value could be anything, so we need to convert
            if cell_value != value:
                continue

            self.load_row(row)
            return

        # If we do not find the user, show a dialog box
        messagebox.showinfo(message='No data found')

    def data_modified(self, field, row, index):
        self.database = xlrd.open_workbook(self.db_path).sheet_by_index(0)
        value = field.get()

        db_row = self.database.row(row)
        cell_value = db_row[index].value

        # Convert floats to ints to remove trailing decimal places
        if type(cell_value) is float:
            cell_value = int(cell_value)

        # Entry value is a str, cell value could be anything, so we need to convert
        if str(cell_value) != value:
            return True

        return False

    def load_row(self, row):
        row_data = self.database.row(row)

        zip_value = row_data[ZIP_INDEX].value

        # Python cannot convert empty strings to ints

        if zip_value is not '' and type(zip_value) is float:
            zip_value = int(zip_value)

        self.set_text(self.donor_field, row_data[DONOR_INDEX].value)
        self.set_text(self.name_field, row_data[NAME_INDEX].value)
        self.set_text(self.street_field, row_data[STREET_INDEX].value)
        self.set_text(self.city_field, row_data[CITY_INDEX].value)
        self.set_text(self.state_field, row_data[STATE_INDEX].value)
        self.set_text(self.zip_field, zip_value)
        self.set_text(self.phone_field, row_data[PHONE_INDEX].value)
        self.set_text(self.email_field, row_data[EMAIL_INDEX].value)
        self.set_text(self.company_field, row_data[COMPANY_INDEX].value)

        self.loaded_row = row

    def submit(self):
        workbook = openpyxl.load_workbook(self.db_path)
        sheet = workbook.worksheets[0]
        is_addition = True

        # Start with green fill for additions
        color = PatternFill(fgColor=colors.GREEN, fill_type='solid')

        # A row is loaded, edit the data
        if self.loaded_row is not None:
            # Set edit color to yellow
            color = PatternFill(fgColor=colors.YELLOW, fill_type='solid')
            row = self.loaded_row + 1
            is_addition = False

        # No row loaded, add new data
        else:
            row = sheet.max_row + 1

        # Only write and highlight modified values
        if is_addition or self.data_modified(self.donor_field, row - 1, DONOR_INDEX):
            sheet.cell(row, DONOR_INDEX + 1, self.donor_field.get()).fill = color

        if is_addition or self.data_modified(self.name_field, row - 1, NAME_INDEX):
            sheet.cell(row, NAME_INDEX + 1, self.name_field.get()).fill = color

        if is_addition or self.data_modified(self.street_field, row - 1, STREET_INDEX):
            sheet.cell(row, STREET_INDEX + 1, self.street_field.get()).fill = color

        if is_addition or self.data_modified(self.city_field, row - 1, CITY_INDEX):
            sheet.cell(row, CITY_INDEX + 1, self.city_field.get()).fill = color

        if is_addition or self.data_modified(self.state_field, row - 1, STATE_INDEX):
            sheet.cell(row, STATE_INDEX + 1, self.state_field.get()).fill = color

        if is_addition or self.data_modified(self.zip_field, row - 1, ZIP_INDEX):
            sheet.cell(row, ZIP_INDEX + 1, self.zip_field.get()).fill = color

        if is_addition or self.data_modified(self.phone_field, row - 1, PHONE_INDEX):
            # Automatically apply formatting to 10-digit US numbers
            phone_number = unformat_phone_number(self.phone_field.get())  # Strip existing formatting
            phone_number = format_phone_number(phone_number)  # Apply new formatting
            sheet.cell(row, PHONE_INDEX + 1, phone_number).fill = color

        if is_addition or self.data_modified(self.email_field, row - 1, EMAIL_INDEX):
            sheet.cell(row, EMAIL_INDEX + 1, self.email_field.get()).fill = color

        if is_addition or self.data_modified(self.company_field, row - 1, COMPANY_INDEX):
            sheet.cell(row, COMPANY_INDEX + 1, self.company_field.get()).fill = color

        try:
            workbook.save(self.db_path)
        except PermissionError:
            messagebox.showerror('Error Saving File', 'Please close Excel to modify the database')
            return
        except:
            messagebox.showerror('Error Saving File', 'An unknown error occurred while trying to save changes')
            return

        self.clear()

        action = 'added' if is_addition else 'edited'
        messagebox.showinfo('Success', f'Data successfully {action}')

    def set_text(self, field, text):
        # Clear existing text
        field.delete(0, tk.END)
        # Insert new text
        field.insert(0, text)

    def confirm_clear(self):
        confirm = messagebox.askyesno('Clear Data', 'Are you sure you want to discard the data?')

        if confirm:
            self.clear()

    def clear(self):
        self.donor_field.delete(0, tk.END)
        self.name_field.delete(0, tk.END)
        self.street_field.delete(0, tk.END)
        self.city_field.delete(0, tk.END)
        self.state_field.delete(0, tk.END)
        self.zip_field.delete(0, tk.END)
        self.phone_field.delete(0, tk.END)
        self.email_field.delete(0, tk.END)
        self.email_field.delete(0,tk.END)
        self.company_field.delete(0, tk.END)

        self.loaded_row = None
        self.phone_field.focus()


if __name__ == '__main__':
    root = tk.Tk()

    main = SowersFund(root)

    root.mainloop()
